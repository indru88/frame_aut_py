import openpyxl

class FuncionesExcel():
    def __init__(self, driver):
        self.driver = driver

    def get_row_count(self, path, sheet_name):
        wb = openpyxl.load_workbook(path)
        sheet = wb[sheet_name]
        return sheet.max_row

    def get_column_count(self, path, sheet_name):
        wb = openpyxl.load_workbook(path)
        sheet = wb[sheet_name]
        return sheet.max_column

    def read_data(self, path, sheet_name, row_num, column_num):
        wb = openpyxl.load_workbook(path)
        sheet = wb[sheet_name]
        return sheet.cell(row=row_num, column=column_num).value
    
